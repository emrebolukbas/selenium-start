from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pytest
import openpyxl
from constants import globalConstants as c


"""Bir önceki ödevde yazdığınız tüm testleri PyTest uyumlu hale getiriniz.
Kullandığımız SauceDemo sitesinde kendi belirlediğiniz en az "3" test case daha yazınız.
En az 1 testiniz parametrize fonksiyonu ile en az 3 farklı veriyle test edilmelidir. """

#pytest koşturabilmek için class isimlerine büyük T ile Test_ fonksiyonlara ise küçük t ile test_ isimlendirerek başla yoksa tanımlamaz !!!
class Test_sauceDemoLogin:
    def setup_method(self): #her test başlangıcında çalışacak fonk.
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() #ekranı büyütür

    def teardown_metahod(self): #her testin bitiminde çalışacak fonk.
        self.driver.quit()

    def test_username_password_empty(self):
         loginButtton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
         loginButtton.click()
         errorMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.ERROR_MESSAGE_XPATH)))
         assert errorMessage.text == "Epic sadface: Username is required"

    def test_password_empty_pass(self):
        userName = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        userName.send_keys("standard_user")
        loginButtton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID,c.LOGIN_BUTTON)))
        loginButtton.click()
        errorMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == "Epic sadface: Password is required"
          
    def test_invalid_login(self):
        userName = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        userName.send_keys("locked_out_user")
        password = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        password.send_keys("secret_sauce")
        loginButtton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButtton.click()
        errorMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."


    def test_product_control(self):
        userName = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        userName.send_keys("standard_user")
        password = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        password.send_keys("secret_sauce")
        loginButtton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButtton.click()
        current_url = self.driver.current_url
        expected_url = c.INVENTORY_URL
        if (expected_url == current_url ):
            firstTest = True
        else:
            firstTest = False
        products = WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.CLASS_NAME, c.ALL_PRODUCT_CONTROL)))
        secondTest = len(products) == 6
        assert firstTest == True and secondTest == True 

#1. case = ürün sepete ekleniyor mu?
#2. case = sepete doğru ürünler gidiyor mu ?
#3. case = alışverişi tamamla        
class Test_sauceDemoProducts:
     #prefix => test_
    def setup_method(self): #her test başlangıcında çalışacak fonk.
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() #ekranı büyütür

    def teardown_metahod(self): #her testin bitiminde çalışacak fonk.
        self.driver.quit()


    def getData():
        excel = openpyxl.load_workbook("data/accepted_Usernames.xlsx")
        sheet = excel["Sheet1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))

        return data #[("standard_user", "secret_sauce"),("error_user","secret_sauce")]
    
    #1. case = ürün sepete ekleniyor mu?
    @pytest.mark.parametrize("username_param, password_param", getData() )
    def test_added_Products(self,username_param,password_param):
        userName = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        userName.send_keys(username_param)
        password = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        password.send_keys(password_param)
        loginButtton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButtton.click()
        current_url = self.driver.current_url
        expected_url = c.INVENTORY_URL
        if (expected_url == current_url ):
            firstTest = True
        else:
            firstTest = False
        addToCart1 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID,"add-to-cart-sauce-labs-backpack")))
        addToCart1.click()
        cart = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.CLASS_NAME, c.CART_TEXT)))
        if (cart.text == "1"):
            pass
        else:
            print("product not added")
        addToCart2 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID,"add-to-cart-sauce-labs-bike-light")))
        addToCart2.click()
        assert cart.text == "2"
    #2. case = sepete doğru ürünler gidiyor mu ?
    def test_areCorrect_Products(self):
        userName = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        userName.send_keys("standard_user")
        password = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        password.send_keys("secret_sauce")
        loginButtton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButtton.click()
        current_url = self.driver.current_url
        expected_url = c.INVENTORY_URL
        if (expected_url == current_url ):
            firstTest = True
        else:
            firstTest = False
        addToCart1 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.ADD_TO_CART_1)))
        addToCart1.click()
        addToCart2 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.ADD_TO_CART_2)))
        addToCart2.click()
        firstProduct = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.FIRST_PRODUCT))).text
        secondProduct = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH,c.SECOND_PRODUCT))).text
        cart = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.CLASS_NAME, c.CART)))
        cart.click()
        product_control_1 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.FIRST_PRODUCT))).text
        product_control_2 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.SECOND_PRODUCT))).text
        assert (firstProduct and product_control_1 == "Sauce Labs Backpack") and (secondProduct and product_control_2 == "Sauce Labs Bike Light")
        
    #3. case = alışverişi tamamla
    def test_completeShopping(self):
        userName = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        userName.send_keys("standard_user")
        password = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        password.send_keys("secret_sauce")
        loginButtton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButtton.click()
        current_url = self.driver.current_url
        expected_url = c.INVENTORY_URL
        if (expected_url == current_url ):
            firstTest = True
        else:
            firstTest = False
        addToCart1 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.ADD_TO_CART_1)))
        addToCart1.click()
        addToCart2 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.ADD_TO_CART_2)))
        addToCart2.click()
        goCart = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.CLASS_NAME, c.CART)))
        goCart.click()
        Checkout = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.CHECKOUT)))
        Checkout.click()
        fistName = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.FIRST_NAME)))
        fistName.send_keys("Emre")
        lastName = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LAST_NAME)))
        lastName.send_keys("Bolukbas")
        postalCode = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.POSTAL_CODE)))
        postalCode.send_keys("12345")
        Continue = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.CONTINUE)))
        Continue.click()
        current_url = self.driver.current_url
        expected_url = c.CHECKOUT_STEP_URL
        if ( expected_url == current_url):
            secondTest = True
        else:
            secondTest = False
        finishShopping = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID,c.FINISH)))
        finishShopping.click()
        completeHeader = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.CLASS_NAME, c.COMPLETE_HEADER))).text
        assert completeHeader == "Thank you for your order!"


             
"""          
testLogin = sauceDemoLogin()
testLogin.username_password_empty()
testLogin.password_empty_pass()
testLogin.invalid_login()
testLogin.product_control() """

#testProducts = sauceDemoProducts()
#testProducts.added_Products()
#testProducts.areCorrect_Products()
#testProducts.completeShopping()